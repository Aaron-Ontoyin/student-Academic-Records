import sys
import os
import openpyxl as xl


def option_6():
    if len(os.listdir("./files/students")) == 0:
        return 0
    print("6. Delete a student")


def option_7_8():
    if len(os.listdir("./files/courses")) == 0:
        return 0
    print("7. Delete a course")
    print("8. See available courses")


def delete_st_files():
    with open("st_file.txt") as st_file:
        index_num = st_file.read()
    os.remove("./files/students/" + index_num.rstrip() + ".xlsx")


def delete_cr_files():
    with open("cr_file.txt") as cr_file:
        cr_code = cr_file.read()
    os.remove("./files/courses/" + cr_code.rstrip() + ".xlsx")


def delete_student():
    index_num = input("Enter index number of Student to delete: ")

    # create a file named as index_num value to use in confirming
    # second index_num input at cpp
    with open("confirm_index_num.txt", "a") as confirm_index_num: 
        confirm_index_num.write(index_num)

    st_file = index_num + ".xlsx"
    if st_file not in os.listdir("./files/students"):
        print(f"No student exists with index number[{index_num}]!!")
        return -1
    
    sheet = xl.load_workbook("./files/students/"+st_file).active
    student_name = sheet["c1"].value

    print(f"Processing {student_name}'s data...")
    courses_enrolled = []
    row_num = 8
    while sheet["a"+str(row_num)].value is not None:
        courses_enrolled.append(sheet["a"+str(row_num)].value) # Put course codes into courses enrolled
        row_num += 1

    with open("courses.txt", "a") as courses: # store course codes in courses.txt for cpp
        for i in courses_enrolled:
            courses.write(i)
            courses.write("\n")


def delete_course():
    cr_code = input("Enter code of the course you want to delete(Note: It's case sensitive for security): ")

    # create a file named cr_code value to use in confirming
    # second cr_code input at cpp
    with open("confirm_cr_code.txt", "a") as confirm_cr_code: 
        confirm_cr_code.write(cr_code)

    cr_file = cr_code + ".xlsx"
    if cr_file not in os.listdir("./files/courses"):
        print(f"No course with code {cr_code} exist!!")
        return -1
    
    sheet = xl.load_workbook("./files/courses/"+cr_file).active
    cr_name = sheet["d1"].value

    print(f"Processing {cr_name}'s data...")
    students_enrolled = []
    row_num = 15
    while sheet["a"+str(row_num)].value is not None:
        students_enrolled.append(sheet["a"+str(row_num)].value) # Put course codes into courses enrolled
        row_num += 1

    with open("students.txt", "a") as students: # store course codes in courses.txt for cpp
        for i in students_enrolled:
            students.write(i)
            students.write("\n")


def numbers():
    # The tries are saparate to avoid errors incase either
    # of students ot courses files absent with the presence
    # of the other
    try:
        os.listdir("./files/courses")
    except:
        os.makedirs("./files/courses")
        
    try:
        os.listdir("./files/students")
    except:
        os.makedirs("./files/students")
        
    print("We currently have " +
        str(len(os.listdir("./files/courses"))) + " Course(s)📚 and")
    print(str(len(os.listdir("./files/students"))) + " Student(s)👨‍🎓👩‍🎓")
    print("")


def display_available_courses():
    for i in os.listdir("./files/courses"):
        courses = i
        sheet = xl.load_workbook("./files/courses/"+ i).active
        course_name = "❖  " + str(sheet["d1"].value)
        print(course_name)


def cr_course():
    with open("cr_details.txt", "r") as file:
        lines = file.readlines()
        course_name = lines[0].strip()
        course_instructor_name = lines[1].strip()
        course_code = lines[2].strip()
        expected_credit_hrs = int(lines[3].strip())

    try:    
        for i in os.listdir("./files/courses"):
            if str(course_code.rstrip()).lower() == str(i[:(len(i)-5)].rstrip()).lower():
                print("Course code already taken❌")
                with open("./cr_code_taken.txt", "a") as cr_code_taken:
                    pass
                return -1
    except:
        pass
    
    course_file = course_code.strip() + ".xlsx"

    wb = xl.Workbook()
    sheet = wb.active


    sheet.merge_cells("a1:b1")
    sheet.merge_cells("a2:b2")
    sheet.merge_cells("a3:b3")
    sheet.merge_cells("d1:g1")
    sheet.merge_cells("d2:g2")
    sheet.merge_cells("d3:e3")
    sheet.merge_cells("a5:c5")
    sheet.merge_cells("a6:c6")
    sheet.merge_cells("a8:b8")
    sheet.merge_cells("a11:b11")
    sheet.merge_cells("a13:c13")
    sheet.merge_cells("b14:e14")

    course_name_cell = sheet["a1"]
    course_instructor_name_cell = sheet["a2"]
    course_code_cell = sheet["a3"]
    number_of_students_ever_enrolled_cell = sheet["a5"]
    number_of_students_currently_enrolled_cell = sheet["a6"]
    number_of_recorded_cell = sheet["a8"]
    quizes_cell = sheet["c8"]
    assignments_cell = sheet["c9"]
    spent_hours_cell = sheet["c10"]
    expected_hours_cell = sheet["a11"]
    students_enrolled_cell = sheet["a13"]

    course_name_value_cell = sheet["d1"]
    course_instructor_name_value_cell = sheet["d2"]
    course_code_value_cell = sheet["d3"]
    number_of_students_ever_enrolled_value_cell = sheet["d5"]
    number_of_students_currently_enrolled_value_cell = sheet["d6"]
    quizes_value_cell = sheet["d8"]
    assignments_value_cell = sheet["d9"]
    spent_hours_value_cell = sheet["d10"]
    expected_hours_value_cell = sheet["d11"]

    course_name_cell.value = "COURSE NAME:"
    course_instructor_name_cell.value = "COURSE INSTRUCTOR'S NAME:"
    course_code_cell.value = "COURSE CODE:"
    number_of_students_ever_enrolled_cell.value = "No. of students ever enrolled:"
    number_of_students_currently_enrolled_cell.value = "No. of students currently enrolled:"
    number_of_recorded_cell.value = "Number of Recorded:"
    quizes_cell.value = "Quizes:"
    assignments_cell.value = "Assignments:"
    spent_hours_cell.value = "Hours:"
    expected_hours_cell.value = "Expected hours:"
    students_enrolled_cell.value = "LIST OF ALL ENROLLED STUDENTS"

    course_name_value_cell.value = course_name
    course_instructor_name_value_cell.value = course_instructor_name
    course_code_value_cell.value = course_code
    number_of_students_ever_enrolled_value_cell.value = 0
    number_of_students_currently_enrolled_value_cell.value = 0
    quizes_value_cell.value = 0
    assignments_value_cell.value = 0
    spent_hours_value_cell.value = 0
    expected_hours_value_cell.value = expected_credit_hrs
    
    # I now stop using variables but just hard write the 
    # constant cell values
    sheet["a14"].value = "Idex No."
    sheet["b14"].value = "Student Name"
    sheet["f14"].value = "Quizes Mrks"
    sheet["g14"].value = "Assignments Mrks"
    sheet["h14"].value = "Attendences Mrks"
    sheet["i14"].value = "Exam Mrk"
    sheet["j14"].value = "Total mrk(%)"
    sheet["k14"].value = "Grade"
  
    wb.save("./files/courses/" + course_file)
    
    success = "$$$$🌈 Course succesfully created with code " + course_code
    print(success)


def rg_student():
    with open("st_details.txt", "r") as file:
        lines = file.readlines()
        student_name = lines[0].strip()
        index_num = lines[1].strip()
        department = lines[2].strip()

    try:
        for i in os.listdir("./files/students"):
            if str(index_num.rstrip()).lower() == str(i[:(len(i)-5)].rstrip()).lower():
                print("Index number already taken❌")
                with open("./index_num_taken.txt", "a") as index_num_taken:
                    pass
                return -1
    except:
        pass
    
    st_file = index_num.strip() + ".xlsx"

    wb = xl.Workbook()
    sheet = wb.active

    sheet.merge_cells("a2:b2")
    sheet.merge_cells("a3:b3")
    sheet.merge_cells("a5:b5")
    sheet.merge_cells("c2:d2")
    sheet.merge_cells("c1:g1")
    sheet.merge_cells("c3:g3")
    sheet.merge_cells("j6:k6")

    sheet["a1"].value = "NAME:"
    sheet["a2"].value = "INDEX NUM:"
    sheet["a3"].value = "DEPARTMENT:"
    sheet["c1"].value = student_name
    sheet["c2"].value = index_num
    sheet["c3"].value = department
    sheet["a5"].value = "COURSES ENROLLED:"
    sheet["a6"].value = "Cr code"
    sheet["b6"].value = "Cr name"
    sheet["f6"].value = "Lecturer"
    sheet["j6"].value = "Marks scored in:"
    sheet["i7"].value = "Quizes"
    sheet["j7"].value = "Assignments"
    sheet["k7"].value = "Attendance"
    sheet["l7"].value = "Exam"
    sheet["M7"].value = "Total mrk"
    sheet["N7"].value = "Grade"

    wb.save("./files/students/" + st_file)
    print(
        f"$$$$🌈 Student succesfully registered with index number {index_num}")


def enroll_student():
    with open("st_and_cr_details.txt", "r") as file:
        lines = file.readlines()
        index_num = lines[0].strip()
        cr_code = lines[1].strip()

    st_wb = xl.load_workbook("./files/students/" + index_num + ".xlsx")
    cr_wb = xl.load_workbook("./files/courses/" + cr_code + ".xlsx")

    cr_free_row = 15
    cr_sheet = cr_wb.active
    st_free_row = 8
    st_sheet = st_wb.active

    while True:
        try:
            if int(cr_sheet["a"+str(cr_free_row)].value) == int(index_num):
                print("Student already enrolled in this course🛑")
                return 0
            elif cr_free_row > 10000:
                break
            cr_free_row += 1
        except:
            break

    cr_free_row = 15  # reassign cr_free_row for registration

    while cr_sheet["a"+str(cr_free_row)].value is not None:
        cr_free_row += 1

    while st_sheet["a"+str(st_free_row)].value is not None:
        st_free_row += 1

    # NOW WORK ON COURSE SHEET
    cr_sheet["a"+str(cr_free_row)].value = index_num

    # Set grade and total mark cells formula
    cr_free_row = str(cr_free_row) #Convert cr_free_row to a str fo rinserting the formulas

    totalMark = f"=((f{cr_free_row}*2/3) + (g{cr_free_row}/3) + (h{cr_free_row}*10/d11) +(i{cr_free_row}*0.6))"
    cr_sheet["j"+str(cr_free_row)].value = totalMark
    
    grade = f'=IF(J{cr_free_row}=0, "**", IF(J{cr_free_row}>=80, "A", IF(J{cr_free_row}>=70, "B", IF(J{cr_free_row}>=60, "C", IF(J{cr_free_row}>=50, "D", "FAIL")))))'
    cr_sheet["k"+str(cr_free_row)].value = grade

    cr_sheet.merge_cells(f"b{cr_free_row}:e{cr_free_row}")

    cr_free_row = int(cr_free_row) # Convert cr_free_row back to into after inserting formulas

    cr_sheet["b"+str(cr_free_row)].value = st_sheet["c1"].value.strip()  # Student name
    cr_sheet["d5"].value += 1
    cr_sheet["d6"].value += 1

    # NOW WORK ON STUDENT SHEET
    st_sheet.merge_cells(f"b{st_free_row}:e{st_free_row}")
    st_sheet.merge_cells(f"f{st_free_row}:h{st_free_row}")

    attendanceValue = cr_sheet["d11"].value
    st_sheet["a"+str(st_free_row)].value = cr_code
    st_sheet["b"+str(st_free_row)].value = cr_sheet["d1"].value  # course name
    st_sheet["f"+str(st_free_row)].value = cr_sheet["d2"].value # course instructor's name

    # set formula for total marks
    totalMark = f"=((i{st_free_row}*2/3) + (j{st_free_row}/3) + (k{st_free_row}*10/{attendanceValue}) +(l{st_free_row}*0.6))"
    st_sheet["m"+str(st_free_row)].value = totalMark

    # set formula for Grade
    grade = f'=IF(M{st_free_row}=0, "**", IF(M{st_free_row}>=80, "A", IF(M{st_free_row}>=70, "B", IF(M{st_free_row}>=60, "C", IF(M{st_free_row}>=50, "D", "FAIL")))))'
    st_sheet["n"+str(st_free_row)].value = grade

    st_wb.save("./files/students/" + index_num + ".xlsx")
    cr_wb.save("./files/courses/" + cr_code + ".xlsx")

    print("🌈Succesfully enrolled a student with:")
    print(f"Index number '{index_num}' to course with:")
    print(f"Course code '{cr_code}'")


def unenroll_student():
    with open("st_and_cr_details.txt", "r") as file:
        lines = file.readlines()
        index_num = lines[0].strip()
        cr_code = lines[1].strip()

    st_wb = xl.load_workbook("./files/students/" + index_num + ".xlsx")
    cr_wb = xl.load_workbook("./files/courses/" + cr_code + ".xlsx")

    cr_sheet = cr_wb.active
    st_sheet = st_wb.active
    
    st_name = st_sheet["c1"].value
    print(f"Are you sure you want to unenroll {st_name} from {cr_code}?")
    confirm = input("[Enter y/Y for yes or any char for no]➤➤ ")
    if confirm.lower() != "y":
        print("Aborting Unenrollment...")
        print("Arboted🚫")
        with open("./aborted_unenrollment.txt", "a") as aborted_unenrollment:
            pass
        return 0
   
    print("Unenrolling...")
  
    cr_sheet["d6"].value -= 1

    # look for row with student num in course sheet
    row_with_student = 15  # 15th row contains first student in course sheet
    try:
        while int(cr_sheet["a"+str(row_with_student)].value) != int(index_num):
            row_with_student += 1
    except:
        print("The student is not enrolled")
        return 0

    cr_sheet.delete_rows(row_with_student)

    # look for course code in student sheet
    row_with_course = 8  # 8th row contains first course in student sheet
    while (st_sheet["a"+str(row_with_course)].value).lower() != cr_code.lower():  # d3 contains cr_code
        row_with_course += 1

    st_sheet.delete_rows(row_with_course)

    st_wb.save("./files/students/" + index_num + ".xlsx")
    cr_wb.save("./files/courses/" + cr_code + ".xlsx")
    print("Finished updating worksheets. Now deleting student's records...")

    print("🌈Succesfully unenrolled a student with:")
    print(f"Index number '{index_num}' from course with:")
    print(f"Course code '{cr_code}'")


def deleting_unenrollment():
    with open("st_and_cr_details.txt", "r") as file:
        lines = file.readlines()
        index_num = lines[0].strip()
        cr_code = lines[1].strip()

    st_wb = xl.load_workbook("./files/students/" + index_num + ".xlsx")
    cr_wb = xl.load_workbook("./files/courses/" + cr_code + ".xlsx")

    cr_sheet = cr_wb.active
    st_sheet = st_wb.active
    
    st_name = st_sheet["c1"].value
    
    cr_sheet["d6"].value -= 1

    # look for row with student num in course sheet
    row_with_student = 15  # 15th row contains first student in course sheet
    while int(cr_sheet["a"+str(row_with_student)].value) != int(index_num):
        row_with_student += 1

    cr_sheet.delete_rows(row_with_student)

    # look for course code in student sheet
    row_with_course = 8  # 8th row contains first course in student sheet
    while (st_sheet["a"+str(row_with_course)].value).lower() != cr_code.lower():  # d3 contains cr_code
        row_with_course += 1

    st_sheet.delete_rows(row_with_course)

    st_wb.save("./files/students/" + index_num + ".xlsx") # Yes, I still save it even though I'll delete it
    cr_wb.save("./files/courses/" + cr_code + ".xlsx")


def record_marks():
    with open("test_and_course_details.txt") as test_cr:
        lines = test_cr.readlines()
        cr_code = lines[0].rstrip()
        index_num = lines[1].rstrip()
        test = lines[2].rstrip()

    st_wb = xl.load_workbook("./files/students/" + index_num + ".xlsx")
    cr_wb = xl.load_workbook("./files/courses/" + cr_code + ".xlsx")
    
    cr_sheet = cr_wb.active
    st_sheet = st_wb.active

    
    st_name = st_sheet["c1"].value
    cr_name = cr_sheet["d1"].value

    st_first_row = 8
    while True:
        if st_sheet["a"+str(st_first_row)].value == cr_code:
            break
        elif st_first_row > 200: # Student assumed to have <= 200 courses
            print(st_name + " not enrolled in " +  cr_name + "🛑")
            return 0
        st_first_row += 1
    
    cr_first_row = 15
    while True:
        if int(cr_sheet["a"+str(cr_first_row)].value) == int(index_num):
            break
        cr_first_row += 1

    # Student and course fisrt rows could be modified by the while loop

    testCopy = test   # Copy test for knowing the columes in st and cr sheets
   # Eg Quiz column in st sheet != in cr sheet

    #Record for student
    if testCopy == "Quiz":
        test = "i"
    elif testCopy == "Assingment":
        test = "j"
    elif testCopy == "Attendance":
        test = "k"
    elif testCopy == "Exam":
        test = "l"

    mark = input(f"Enter {st_name}'s mark for the {cr_name} {testCopy}: ")
    st_sheet[test+str(st_first_row)].value = mark

    #Record for course
    if testCopy == "Quiz":
        test = "f"
    elif testCopy == "Assingment":
        test = "g"
    elif testCopy == "Attendance":
        test = "h"
    elif testCopy == "Exam":
        test = "i"

    cr_sheet[test+str(cr_first_row)].value = mark

    st_wb.save("./files/students/" + index_num + ".xlsx")
    cr_wb.save("./files/courses/" + cr_code + ".xlsx")

    print("Successfully recorded *&&&*")

if __name__ == '__main__':
    globals()[sys.argv[1]]()
